"""
Money Tracker Pro - Main Window
World-Class UI/UX Design Implementation
"""

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QListWidget, QListWidgetItem,
    QPushButton, QFrame, QStackedWidget, QSplitter,
    QCheckBox, QApplication, QMenu, QGraphicsDropShadowEffect,
    QDateEdit, QDialog, QScrollBar
)
from PyQt6.QtCore import Qt, QTimer, QPropertyAnimation, QEasingCurve, pyqtSignal, QSize, QSettings, QDate
from PyQt6.QtGui import QFont, QColor, QIcon, QAction, QKeySequence, QShortcut, QWheelEvent

# 保持原有导入，确保逻辑兼容
from .styles import MAIN_STYLESHEET, get_theme_stylesheet, DARK_THEME
from .animated_number import AnimatedNumberLabel, AnimatedStatLabel, SelectionStatsWidget
from .transaction_dialog import TransactionDialog
from .module_dialog import ModuleDialog
from .settings_dialog import SettingsDialog
from .summary_panel import SummaryPanel
from common.database import Database

import csv
import os
import subprocess
from datetime import datetime

# --- 顶级 UI 样式配置 (内嵌) ---
MODERN_STYLES = """
    /* 全局字体与重置 */
    * {
        font-family: "SF Pro Display", "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
        outline: none;
    }
    
    /* 主窗口背景 - 深空灰黑 */
    QMainWindow, QWidget#centralWidget {
        background-color: #0F0F12; 
    }

    /* 滚动条美化 - 极简风格 */
    QScrollBar:vertical {
        border: none;
        background: transparent;
        width: 6px;
        margin: 0;
    }
    QScrollBar::handle:vertical {
        background: #333338;
        min-height: 20px;
        border-radius: 3px;
    }
    QScrollBar::handle:vertical:hover {
        background: #4D4D55;
    }
    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
        height: 0px;
    }

    /* QListWidget 基础清理 */
    QListWidget {
        background: transparent;
        border: none;
        outline: none;
    }
    QListWidget::item {
        background: transparent;
        padding: 4px;
        border: none;
    }
    QListWidget::item:selected {
        background: transparent;
        border: none;
    }
    QListWidget::item:hover {
        background: transparent;
    }

    /* Tooltip */
    QToolTip {
        background-color: #1F1F24;
        color: #E0E0E0;
        border: 1px solid #333338;
        border-radius: 4px;
        padding: 6px;
        font-size: 12px;
    }
"""

# --- 辅助动效函数 ---
def apply_shadow(widget, blur=15, x=0, y=4, alpha=80):
    """为组件添加高级阴影"""
    from .effects import glow_shadow
    glow_shadow(widget, "#000000", blur, alpha, x, y)


class SmoothListWidget(QListWidget):
    """A list widget with dampened scroll speed."""
    def wheelEvent(self, event: QWheelEvent):
        dampened = event.angleDelta() / 3
        new_event = QWheelEvent(
            event.position(), event.globalPosition(),
            event.pixelDelta(), dampened,
            event.buttons(), event.modifiers(),
            event.phase(), event.inverted()
        )
        super().wheelEvent(new_event)


class ActionButton(QFrame):
    """
    Custom action button with Python-driven hover.
    Uses enterEvent/leaveEvent instead of CSS :hover to avoid
    the ghost double-render bug on Windows QPushButton.
    """
    clicked = pyqtSignal()

    def __init__(self, text: str, color: str, bg_color: str, parent=None):
        super().__init__(parent)
        self._color   = color
        self._bg_norm = bg_color
        self._bg_hover = self._lighten(bg_color)
        self._is_hover = False

        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setFixedHeight(46)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        self._lbl = QLabel(text)
        self._lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._lbl.setStyleSheet(
            f"font-size: 14px; font-weight: 700; color: {color};"
            f" background: transparent; border: none; pointer-events: none;"
        )
        lay.addWidget(self._lbl)
        self._apply_style(False)

    @staticmethod
    def _lighten(rgba_str: str) -> str:
        """Return a slightly brighter version for hover."""
        # Just bump opacity — insert a new rgba with higher alpha
        return rgba_str.replace("0.10", "0.22").replace("0.08", "0.20")

    def _apply_style(self, hovered: bool):
        bg = self._bg_hover if hovered else self._bg_norm
        border_alpha = "80" if hovered else "44"
        self.setStyleSheet(f"""
            ActionButton {{
                background: {bg};
                border: 1px solid {self._color}{border_alpha};
                border-radius: 12px;
            }}
        """)

    def enterEvent(self, event):
        self._apply_style(True)
        super().enterEvent(event)

    def leaveEvent(self, event):
        self._apply_style(False)
        super().leaveEvent(event)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit()
        super().mousePressEvent(event)


class TransactionItemWidget(QWidget):
    """
    World-Class Transaction Card
    Features: Glassmorphism background, semantic colors, smooth hover state.
    """
    toggled = pyqtSignal(int, bool)
    
    def __init__(self, trans_id: int, amount: float, trans_type: str, 
                 source: str, is_selected: bool, date_str: str, 
                 show_checkbox: bool = False, module_info: tuple = None, parent=None):
        super().__init__(parent)
        self.trans_id = trans_id
        self.show_checkbox = show_checkbox
        self.is_selected = is_selected
        self.is_income = (trans_type == "income")
        
        from .styles import THEME
        self.theme = THEME
        
        self.card_frame = QFrame(self)
        self.card_frame.setObjectName("cardFrame")
        
        self.main_layout = QHBoxLayout(self)
        self.main_layout.setContentsMargins(6, 4, 6, 4)
        self.main_layout.addWidget(self.card_frame)
        
        self.card_layout = QHBoxLayout(self.card_frame)
        self.card_layout.setContentsMargins(16, 12, 20, 12)
        self.card_layout.setSpacing(16)
        
        self._setup_ui(amount, trans_type, source, date_str, module_info)
        self._update_style(False)
        # Only track mouse on the outer widget, NOT on card_frame
        # Double tracking on both causes the double-render ghost effect
        self.setMouseTracking(True)

    def _setup_ui(self, amount, trans_type, source, date_str, module_info):
        self.checkbox = QCheckBox()
        self.checkbox.setChecked(self.is_selected)
        self.checkbox.setCursor(Qt.CursorShape.PointingHandCursor)
        self.checkbox.stateChanged.connect(self._on_toggle)
        self.checkbox.setStyleSheet(f"""
            QCheckBox::indicator {{ width: 18px; height: 18px; border-radius: 5px; border: 2px solid {self.theme['border_bright']}; background: transparent; }}
            QCheckBox::indicator:checked {{ background: {self.theme['accent']}; border-color: {self.theme['accent']}; }}
            QCheckBox::indicator:hover {{ border-color: {self.theme['accent_deep']}; }}
        """)
        self.card_layout.addWidget(self.checkbox)
        self.checkbox.setVisible(self.show_checkbox)
        
        accent = self.theme["income"] if self.is_income else self.theme["expense"]
        accent_bg = self.theme["income_soft"] if self.is_income else self.theme["expense_soft"]
        icon_char = "↗" if self.is_income else "↙"
        
        icon_box = QLabel(icon_char)
        icon_box.setFixedSize(42, 42)
        icon_box.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon_box.setStyleSheet(f"""
            QLabel {{
                color: {accent};
                background: {accent_bg};
                border: 1px solid {accent}33;
                border-radius: 12px;
                font-size: 18px;
                font-weight: 900;
            }}
        """)
        self.card_layout.addWidget(icon_box)
        
        info_layout = QVBoxLayout()
        info_layout.setSpacing(4)
        
        source_text = source if source else "未分类"
        lbl_source = QLabel(source_text)
        lbl_source.setStyleSheet(f"font-size:15px; font-weight:700; color:{self.theme['text_p']}; background:transparent;")
        info_layout.addWidget(lbl_source)
        
        meta_text = date_str
        if module_info:
            mod_icon, mod_name = module_info
            meta_text = f"{mod_icon} {mod_name}  ·  {date_str}"
            
        lbl_meta = QLabel(meta_text)
        lbl_meta.setStyleSheet(f"font-size:12px; color:{self.theme['text_t']}; background:transparent;")
        info_layout.addWidget(lbl_meta)
        
        self.card_layout.addLayout(info_layout, 1)
        
        sign = "+" if self.is_income else "−"
        
        amt_layout = QVBoxLayout()
        amt_layout.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        amt_layout.setSpacing(2)
        
        exact_str = f"{amount:,.2f}"
        val_str = exact_str
        if abs(amount) >= 10_000:
            val_str = f"{amount/10_000:,.2f}万"
            
        lbl_amount = QLabel(f"{sign}¥{val_str}")
        lbl_amount.setAlignment(Qt.AlignmentFlag.AlignRight)
        lbl_amount.setStyleSheet(f"""
            font-size: 16px; font-weight: 800; color: {accent}; 
            letter-spacing: 0.5px; background: transparent;
        """)
        # Always show exact full value on hover (important for amounts >= 10,000)
        lbl_amount.setToolTip(f"{sign}¥{exact_str}")
        lbl_amount.setToolTipDuration(6000)
        
        lbl_currency = QLabel("元")
        lbl_currency.setAlignment(Qt.AlignmentFlag.AlignRight)
        lbl_currency.setStyleSheet(f"font-size: 10px; color: {self.theme['text_t']}; font-weight: 700; background: transparent;")
        
        amt_layout.addWidget(lbl_amount)
        amt_layout.addWidget(lbl_currency)
        self.card_layout.addLayout(amt_layout)

    def _update_style(self, hovered: bool):
        base = self.theme['bg_layer2'] if hovered else self.theme['bg_layer1']
        bord = self.theme['border_bright'] if hovered else self.theme['border']
        if self.checkbox.isChecked():
            base = self.theme['accent_ultra']
            bord = self.theme['border_gold']
            
        self.card_frame.setStyleSheet(f"""
            QFrame#cardFrame {{
                background: {base};
                border-radius: 14px;
                border: 1px solid {bord};
            }}
        """)
        
    def enterEvent(self, e):
        self._update_style(True)
        super().enterEvent(e)
        
    def leaveEvent(self, e):
        self._update_style(False)
        super().leaveEvent(e)

    def _on_toggle(self, state):
        checked = (state == Qt.CheckState.Checked.value)
        self._update_style(checked)  # Pass current hover state correctly
        self.toggled.emit(self.trans_id, checked)
    
    def set_checkbox_visible(self, visible: bool):
        self.show_checkbox = visible
        self.checkbox.setVisible(visible)
        if not visible:
            self.checkbox.setChecked(False)
            self._update_style(False)


class ModuleContentPanel(QWidget):
    """
    World-Class Ledger Content Panel
    Features: Animated hero balance card, glassmorphism action buttons, smooth transaction list.
    """
    data_changed = pyqtSignal()   # emitted whenever a transaction is added / deleted

    def __init__(self, db: Database, parent=None):
        super().__init__(parent)
        self.db = db
        self.module_id = None
        self.module_name = ""
        
        from .styles import THEME
        self.theme = THEME
        
        self.is_select_mode = False
        self.selected_transactions = set()
        self._raw_transactions = []          # stores fetched transactions for sorting
        self._sort_desc = True               # True = descending (default), False = ascending
        self._filter_type = "all"            # "all" | "income" | "expense"
        self._setup_ui()
        
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 28, 28, 24)
        layout.setSpacing(24)
        
        # ── Hero Header ──────────────────────────────────────────
        header_layout = QHBoxLayout()
        header_layout.setSpacing(16)
        
        self.icon_lbl = QLabel("💰")
        self.icon_lbl.setFixedSize(64, 64)
        self.icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.icon_lbl.setStyleSheet(f"""
            QLabel {{
                background: {self.theme['bg_layer2']};
                border: 1px solid {self.theme['border_bright']};
                border-radius: 20px;
                font-size: 32px;
            }}
        """)
        header_layout.addWidget(self.icon_lbl)
        
        title_vbox = QVBoxLayout()
        title_vbox.setSpacing(4)
        title_vbox.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        
        self.title_lbl = QLabel("未选择账本")
        self.title_lbl.setStyleSheet(f"font-size: 28px; font-weight: 800; color: {self.theme['text_p']}; letter-spacing: 0.5px;")
        title_vbox.addWidget(self.title_lbl)
        
        self.balance_val = AnimatedNumberLabel(0.0)
        self.balance_val.set_auto_color(True)
        # override font size for this specific context
        f = self.balance_val.label.font()
        f.setPointSize(22)
        self.balance_val.label.setFont(f)
        title_vbox.addWidget(self.balance_val)
        
        header_layout.addLayout(title_vbox)
        header_layout.addStretch()
        
        self.btn_multi_select = QPushButton("☑ 管理")
        self.btn_multi_select.setCheckable(True)
        self.btn_multi_select.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_multi_select.setFixedSize(90, 40)
        self.btn_multi_select.setStyleSheet(f"""
            QPushButton {{
                background: {self.theme['bg_layer2']}; color: {self.theme['text_s']};
                border: 1px solid {self.theme['border']}; border-radius: 12px;
                font-size: 14px; font-weight: 700;
            }}
            QPushButton:hover {{ background: {self.theme['bg_layer3']}; color: {self.theme['text_p']}; }}
            QPushButton:checked {{ 
                background: {self.theme['accent_ultra']}; color: {self.theme['accent']};
                border-color: {self.theme['accent']}66; 
            }}
        """)
        self.btn_multi_select.toggled.connect(self._toggle_multi_select)
        header_layout.addWidget(self.btn_multi_select)
        
        layout.addLayout(header_layout)
        
        # ── Stats Summary ────────────────────────────────────────
        stats_frame = QFrame()
        stats_frame.setStyleSheet(f"""
            QFrame {{
                background: {self.theme['bg_layer1']};
                border: 1px solid {self.theme['border']};
                border-radius: 16px;
            }}
        """)
        stats_layout = QHBoxLayout(stats_frame)
        stats_layout.setContentsMargins(24, 16, 24, 16)
        
        self.month_income_lbl = AnimatedStatLabel(0.0, "收 ¥")
        self.month_income_lbl.setStyleSheet(f"font-size: 15px; font-weight: 700; color: {self.theme['income']};")
        stats_layout.addWidget(self.month_income_lbl)
        
        stats_layout.addStretch()
        
        sep = QFrame()
        sep.setFixedWidth(1)
        sep.setStyleSheet(f"background: {self.theme['border']};")
        stats_layout.addWidget(sep)
        
        stats_layout.addStretch()
        
        self.month_expense_lbl = AnimatedStatLabel(0.0, "支 ¥")
        self.month_expense_lbl.setStyleSheet(f"font-size: 15px; font-weight: 700; color: {self.theme['expense']};")
        stats_layout.addWidget(self.month_expense_lbl)
        
        layout.addWidget(stats_frame)
        
        # ── Control Bar ──────────────────────────────────────────
        self.control_bar = QFrame()
        cb_layout = QHBoxLayout(self.control_bar)
        cb_layout.setContentsMargins(4, 0, 4, 0)
        
        self.btn_select_all = QPushButton("全选")
        self.btn_select_all.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_select_all.clicked.connect(self._select_all)
        
        self.btn_delete_selected = QPushButton("删除选中")
        self.btn_delete_selected.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_delete_selected.setStyleSheet(f"""
            QPushButton {{ background: rgba(255,69,58,0.15); color: #FF453A; border: none; font-weight: 700; }}
            QPushButton:hover {{ background: rgba(255,69,58,0.25); }}
        """)
        self.btn_delete_selected.clicked.connect(self._delete_selected)
        
        cb_layout.addWidget(self.btn_select_all)
        cb_layout.addWidget(self.btn_delete_selected)
        cb_layout.addStretch()
        
        # Embedded selection stats
        self.selection_stats = SelectionStatsWidget()
        cb_layout.addWidget(self.selection_stats)
        
        self.control_bar.hide()
        layout.addWidget(self.control_bar)
        
        # ── Income/Expense Action Buttons (ActionButton: Python hover, no CSS :hover) ──
        action_bar = QHBoxLayout()
        action_bar.setSpacing(12)
        
        btn_income = ActionButton(
            "+  记入收入",
            self.theme['income'],
            self.theme['income_soft']
        )
        btn_income.clicked.connect(lambda: self._add_transaction("income"))
        
        btn_expense = ActionButton(
            "−  记录支出",
            self.theme['expense'],
            self.theme['expense_soft']
        )
        btn_expense.clicked.connect(lambda: self._add_transaction("expense"))
        
        action_bar.addWidget(btn_income, 1)
        action_bar.addWidget(btn_expense, 1)
        layout.addLayout(action_bar)
        
        # ── Filter + Sort Control Bar ──────────────────────────────
        ctrl_bar = QHBoxLayout()
        ctrl_bar.setContentsMargins(2, 0, 2, 0)
        ctrl_bar.setSpacing(7)

        # Type filter pills ── 全部 / 收入↗ / 支出↙
        from PyQt6.QtWidgets import QComboBox

        filter_lbl = QLabel("筛选:")
        filter_lbl.setStyleSheet(f"font-size: 12px; color: {self.theme['text_t']}; background: transparent;")
        ctrl_bar.addWidget(filter_lbl)

        self.filter_btn_all     = QPushButton("全部")
        self.filter_btn_income  = QPushButton("收入 ↗")
        self.filter_btn_expense = QPushButton("支出 ↙")
        self._filter_btns = {
            "all":     self.filter_btn_all,
            "income":  self.filter_btn_income,
            "expense": self.filter_btn_expense,
        }
        for key, btn in self._filter_btns.items():
            btn.setFixedHeight(30)
            btn.setCheckable(True)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            ctrl_bar.addWidget(btn)
            btn.clicked.connect(lambda checked, k=key: self._set_filter_type(k))
        self._apply_filter_btn_styles()  # mark "all" as active

        # Thin separator
        sep = QFrame()
        sep.setFixedWidth(1)
        sep.setFixedHeight(18)
        sep.setStyleSheet(f"background: {self.theme['border']}; border: none;")
        ctrl_bar.addWidget(sep)

        # Sort section
        sort_lbl = QLabel("排序:")
        sort_lbl.setStyleSheet(f"font-size: 12px; color: {self.theme['text_t']}; background: transparent;")
        ctrl_bar.addWidget(sort_lbl)

        self.sort_combo = QComboBox()
        self.sort_combo.addItems(["按日期", "按金额"])
        self.sort_combo.setFixedHeight(30)
        self.sort_combo.setFixedWidth(88)
        self.sort_combo.setCursor(Qt.CursorShape.PointingHandCursor)
        self.sort_combo.setStyleSheet(f"""
            QComboBox {{
                background: {self.theme['bg_layer2']};
                border: 1px solid {self.theme['border']};
                border-radius: 8px;
                padding: 0 10px;
                font-size: 12px;
                color: {self.theme['text_s']};
            }}
            QComboBox:hover {{
                border-color: {self.theme['border_bright']};
                color: {self.theme['text_p']};
            }}
            QComboBox::drop-down {{ border: none; width: 20px; }}
            QComboBox QAbstractItemView {{
                background: {self.theme['bg_layer2']};
                border: 1px solid {self.theme['border_bright']};
                border-radius: 6px;
                color: {self.theme['text_p']};
                selection-background-color: {self.theme['accent']}33;
            }}
        """)
        self.sort_combo.currentIndexChanged.connect(lambda _: self._apply_sort())
        ctrl_bar.addWidget(self.sort_combo)

        self.btn_sort_dir = QPushButton("降序 ↓")
        self.btn_sort_dir.setFixedSize(74, 30)
        self.btn_sort_dir.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_sort_dir.setCheckable(True)
        self.btn_sort_dir.setChecked(False)  # False = descending (default)
        self._update_sort_btn_style(False)
        self.btn_sort_dir.clicked.connect(self._toggle_sort_dir)
        ctrl_bar.addWidget(self.btn_sort_dir)

        ctrl_bar.addStretch()
        layout.addLayout(ctrl_bar)

        # ── Smooth Transaction List ────────────────────────
        self.trans_list = SmoothListWidget()
        self.trans_list.setStyleSheet(f"""
            QListWidget {{ background: transparent; border: none; }}
            QListWidget::item {{ padding: 2px 0; }}
        """)
        layout.addWidget(self.trans_list, 1)


    def _update_sort_btn_style(self, is_asc: bool):
        """Style the sort-direction button to reflect current direction."""
        if is_asc:
            label = "升序 ↑"
            bg = f"{self.theme['accent']}22"
            border = f"{self.theme['accent']}66"
            color = self.theme['accent']
        else:
            label = "降序 ↓"
            bg = self.theme['bg_layer2']
            border = self.theme['border']
            color = self.theme['text_s']
        self.btn_sort_dir.setText(label)
        self.btn_sort_dir.setStyleSheet(f"""
            QPushButton {{
                background: {bg};
                border: 1px solid {border};
                border-radius: 8px;
                font-size: 12px; font-weight: 700;
                color: {color};
            }}
            QPushButton:hover {{
                background: {self.theme['bg_layer3']};
                border-color: {self.theme['border_bright']};
                color: {self.theme['text_p']};
            }}
        """)

    def _toggle_sort_dir(self):
        """Toggle asc/desc, update button, and re-render."""
        self._sort_desc = not self._sort_desc
        self._update_sort_btn_style(not self._sort_desc)  # is_asc = not desc
        self._apply_sort()

    def _apply_filter_btn_styles(self):
        """Update pill button styles to reflect the active filter."""
        t = self.theme
        styles = {
            "all": (
                f"rgba(255,184,0,0.18)", t['accent'],
                f"{t['accent']}66",
                t['bg_layer2'], t['text_s'], t['border']
            ),
            "income": (
                f"{t['income_soft']}", t['income'],
                f"{t['income']}66",
                t['bg_layer2'], t['text_s'], t['border']
            ),
            "expense": (
                f"{t['expense_soft']}", t['expense'],
                f"{t['expense']}66",
                t['bg_layer2'], t['text_s'], t['border']
            ),
        }
        for key, btn in self._filter_btns.items():
            active_bg, active_color, active_border, norm_bg, norm_color, norm_border = styles[key]
            is_active = (key == self._filter_type)
            bg     = active_bg     if is_active else norm_bg
            color  = active_color  if is_active else norm_color
            border = active_border if is_active else norm_border
            btn.setChecked(is_active)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {bg};
                    color: {color};
                    border: 1px solid {border};
                    border-radius: 8px;
                    font-size: 12px; font-weight: 700;
                    padding: 0 10px;
                }}
                QPushButton:hover {{
                    background: {active_bg};
                    border-color: {active_border};
                    color: {active_color};
                }}
            """)

    def _set_filter_type(self, key: str):
        """Set the active type filter and re-render."""
        self._filter_type = key
        self._apply_filter_btn_styles()
        self._apply_sort()

    def _apply_sort(self):
        """Filter by type, sort by selected field/direction, and re-render list."""
        if not hasattr(self, '_raw_transactions') or not self._raw_transactions:
            return

        # 1. Type filter
        ft = getattr(self, '_filter_type', 'all')
        if ft == 'income':
            pool = [t for t in self._raw_transactions if t[2] == 'income']
        elif ft == 'expense':
            pool = [t for t in self._raw_transactions if t[2] == 'expense']
        else:
            pool = self._raw_transactions

        # 2. Sort
        sort_by = self.sort_combo.currentIndex()  # 0=date, 1=amount
        if sort_by == 1:
            key_fn = lambda t: t[1]         # amount
        else:
            key_fn = lambda t: t[5] or ""   # created_at date string

        sorted_txns = sorted(pool, key=key_fn, reverse=self._sort_desc)

        # 3. Re-render
        self.trans_list.clear()
        for t in sorted_txns:
            trans_id, amount, trans_type, source, is_selected, created_at = t
            date_str = created_at[:10] if created_at else ""
            widget = TransactionItemWidget(
                trans_id, amount, trans_type, source,
                bool(is_selected), date_str, self.is_select_mode
            )
            widget.toggled.connect(self._on_item_toggled)
            item = QListWidgetItem()
            item.setSizeHint(QSize(0, 88))
            item.setData(Qt.ItemDataRole.UserRole, trans_id)
            self.trans_list.addItem(item)
            self.trans_list.setItemWidget(item, widget)


    def load_module(self, module_id: int, module_name: str, icon: str):
        self.module_id = module_id
        self.module_name = module_name
        
        self.title_lbl.setText(module_name)
        self.icon_lbl.setText(icon)
        
        self.btn_multi_select.setChecked(False)
        self.is_select_mode = False
        self.selected_transactions.clear()
        self.trans_list.clear()
        
        if not module_id:
            return
            
        balance = self.db.get_module_balance(module_id)
        self.balance_val.animate_to(balance)
        
        transactions = self.db.get_transactions_by_module(module_id)
        
        t_inc = sum(t[1] for t in transactions if t[2] == 'income')
        t_exp = sum(t[1] for t in transactions if t[2] == 'expense')
        
        self.month_income_lbl.animate_to(t_inc)
        self.month_expense_lbl.animate_to(t_exp)
        
        # Store for sorting and re-render
        self._raw_transactions = list(transactions)
        self._apply_sort()


    def _toggle_multi_select(self, checked: bool):
        self.is_select_mode = checked
        self.control_bar.setVisible(checked)
        self.selected_transactions.clear()
        self._update_selection_stats()
        
        for i in range(self.trans_list.count()):
            item = self.trans_list.item(i)
            widget = self.trans_list.itemWidget(item)
            widget.set_checkbox_visible(checked)
        
        if not checked:
            self.db.clear_all_selections(self.module_id)
            self.load_module(self.module_id, self.module_name, self.icon_lbl.text())

    def _on_item_toggled(self, trans_id: int, is_selected: bool):
        if is_selected:
            self.selected_transactions.add(trans_id)
        else:
            self.selected_transactions.discard(trans_id)
        self.db.toggle_selection(trans_id) # Update DB selection state
        self._update_selection_stats()
    
    def _update_selection_stats(self):
        count = len(self.selected_transactions)
        if count > 0:
            # get_selected_stats returns (count, income, expense) for currently selected rows in db
            _, t_inc, t_exp = self.db.get_selected_stats(self.module_id)
            self.selection_stats.update_stats(count, t_inc, t_exp)
        else:
            self.selection_stats.update_stats(0, 0.0, 0.0)

    def _select_all(self):
        for i in range(self.trans_list.count()):
            item = self.trans_list.item(i)
            widget = self.trans_list.itemWidget(item)
            if not widget.checkbox.isChecked():
                widget.checkbox.setChecked(True)
        self._update_selection_stats()

    def _delete_selected(self):
        if not self.selected_transactions:
            return
        
        from PyQt6.QtWidgets import QMessageBox
        reply = QMessageBox.question(self, "确认删除", 
                                     f"确定要删除选中的 {len(self.selected_transactions)} 条交易记录吗？",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            for trans_id in list(self.selected_transactions): # Iterate over a copy
                self.db.delete_transaction(trans_id)
            self.selected_transactions.clear()
            self.load_module(self.module_id, self.module_name, self.icon_lbl.text())
            self.btn_multi_select.setChecked(False) # Exit multi-select mode
            self.data_changed.emit()  # notify main window to refresh sidebar + analytics

    def _add_transaction(self, trans_type: str):
        dialog = TransactionDialog(trans_type, self.module_name, self)
        if dialog.exec():
            amount, source = dialog.get_result()
            self.db.add_transaction(amount, trans_type, source, self.module_id)
            self.load_module(self.module_id, self.module_name, self.icon_lbl.text())
            self.data_changed.emit()  # notify main window to refresh sidebar + analytics

    def _on_search(self, keyword: str):
        self.trans_list.clear()
        transactions = self.db.search_transactions(keyword, self.module_id)
        
        from .effects import slide_in_from_bottom
        delay = 0
        for t in transactions:
            trans_id, amount, trans_type, source, is_selected, created_at = t
            date_str = created_at[:10] if created_at else ""
            
            widget = TransactionItemWidget(
                trans_id, amount, trans_type, source, 
                bool(is_selected), date_str, self.is_select_mode
            )
            widget.toggled.connect(self._on_item_toggled)
            
            item = QListWidgetItem()
            item.setSizeHint(QSize(0, 88)) 
            item.setData(Qt.ItemDataRole.UserRole, trans_id)
            
            self.trans_list.addItem(item)
            self.trans_list.setItemWidget(item, widget)
            slide_in_from_bottom(widget, duration=350, distance=20, delay=delay)
            delay = min(delay + 30, 300)

    def _show_context_menu(self, pos):
        item = self.trans_list.itemAt(pos)
        if item:
            trans_id = item.data(Qt.ItemDataRole.UserRole)
            menu = QMenu(self)
            menu.setStyleSheet("""
                QMenu { background: #2C2C2E; border: 1px solid #454545; border-radius: 8px; padding: 4px; }
                QMenu::item { color: #FFF; padding: 6px 20px; border-radius: 4px; }
                QMenu::item:selected { background: #0A84FF; }
            """)
            delete_action = menu.addAction("🗑️ 删除记录")
            action = menu.exec(self.trans_list.mapToGlobal(pos))
            if action == delete_action:
                self.db.delete_transaction(trans_id)
                self.load_module(self.module_id, self.module_name, self.icon_lbl.text())

    def _export_module(self):
        """导出当前账本为 XLSX 表格（openpyxl）"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        settings = QSettings("MoneyTracker", "Shortcuts")
        export_dir = settings.value("export_path", "C:\\")
        
        os.makedirs(export_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = self.module_name.replace(' ', '_').replace('/', '_')
        filename = f"{safe_name}_导出_{timestamp}.xlsx"
        filepath = os.path.join(export_dir, filename)
        
        transactions = self.db.get_transactions_by_module(self.module_id)
        balance = self.db.get_module_balance(self.module_id)
        income, expense = self.db.get_module_stats(self.module_id)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = self.module_name[:31]  # Excel限制31字符
            
            # 样式定义
            title_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
            title_fill = PatternFill(start_color='FF9500', end_color='FF9500', fill_type='solid')
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
            data_font = Font(name='微软雅黑', size=10)
            income_font = Font(name='微软雅黑', size=10, color='FF0000')
            expense_font = Font(name='微软雅黑', size=10, color='009900')
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            thin_border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )
            
            # 概览区域
            summary_data = [
                ("账本名称", self.module_name),
                ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("总余额", f"¥{balance:,.2f}"),
                ("总收入", f"¥{income:,.2f}"),
                ("总支出", f"¥{expense:,.2f}"),
                ("交易笔数", len(transactions)),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                cell_a = ws.cell(row=row_idx, column=1, value=label)
                cell_a.font = title_font
                cell_a.fill = title_fill
                cell_a.alignment = center_align
                cell_b = ws.cell(row=row_idx, column=2, value=value)
                cell_b.font = Font(name='微软雅黑', size=11, bold=True)
                cell_b.alignment = left_align
            
            # 空行
            start_row = len(summary_data) + 2
            
            # 表头
            headers = ["序号", "日期", "类型", "金额", "备注"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            # 数据行
            for i, trans in enumerate(transactions, 1):
                trans_id, amount, trans_type, source, is_selected, created_at = trans
                row = start_row + i
                
                type_str = "收入" if trans_type == "income" else "支出"
                date_str = created_at[:10] if created_at else ""
                signed_amount = amount if trans_type == "income" else -amount
                amt_font = income_font if trans_type == "income" else expense_font
                
                cells = [
                    (1, i),
                    (2, date_str),
                    (3, type_str),
                    (4, round(signed_amount, 2)),
                    (5, source or ""),
                ]
                for col, val in cells:
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.font = amt_font if col == 4 else data_font
                    cell.alignment = center_align if col <= 3 else left_align
                    cell.border = thin_border
                
                # 金额格式
                ws.cell(row=row, column=4).number_format = '#,##0.00'
            
            # 自动列宽
            col_widths = [8, 14, 8, 16, 30]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = w
            
            wb.save(filepath)
            self._show_export_success(filepath, export_dir)
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "导出失败", f"导出时发生错误：\n{str(e)}")
    
    def _show_export_success(self, filepath: str, export_dir: str):
        """显示顶级导出成功对话框"""
        dialog = QDialog(self)
        dialog.setWindowFlags(Qt.WindowType.Dialog | Qt.WindowType.FramelessWindowHint)
        dialog.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        dialog.setFixedSize(400, 300)
        
        container = QFrame(dialog)
        container.setObjectName("exportSuccessCard")
        container.setGeometry(0, 0, 400, 300)
        container.setStyleSheet("""
            QFrame#exportSuccessCard {
                background: #1A1A24;
                border: 1px solid #2A2A3A;
                border-radius: 20px;
            }
        """)
        apply_shadow(container, blur=40, alpha=80)
        
        layout = QVBoxLayout(container)
        layout.setContentsMargins(32, 32, 32, 32)
        layout.setSpacing(16)
        
        icon = QLabel("✅")
        icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon.setStyleSheet("font-size: 48px; background: transparent; border: none;")
        layout.addWidget(icon)
        
        title = QLabel("导出成功！")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("font-size: 22px; font-weight: 800; color: #FFF; letter-spacing: 1px;")
        layout.addWidget(title)
        
        fname = os.path.basename(filepath)
        info = QLabel(f"文件: {fname}")
        info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        info.setStyleSheet("font-size: 12px; color: #888; background: transparent; border: none;")
        info.setWordWrap(True)
        layout.addWidget(info)
        
        layout.addStretch()
        
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(12)
        
        # 修复：正确打开文件所在文件夹并选中文件
        abs_path = os.path.abspath(filepath).replace('/', '\\')
        
        btn_open = QPushButton("📂 打开文件夹")
        btn_open.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_open.setStyleSheet("""
            QPushButton {
                background: #10B981; color: #FFF; border: none;
                border-radius: 12px; min-height: 44px; padding: 0 24px;
                font-weight: 700; font-size: 14px;
            }
            QPushButton:hover { background: #34D399; }
        """)
        btn_open.clicked.connect(lambda checked, p=abs_path: (
            os.startfile(os.path.dirname(p)),
            dialog.accept()
        ))
        
        btn_close = QPushButton("关闭")
        btn_close.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_close.setStyleSheet("""
            QPushButton {
                background: #1A1A24; color: #A0A0B0; border: 1px solid #2A2A3A;
                border-radius: 12px; min-height: 44px; padding: 0 24px;
                font-weight: 600; font-size: 14px;
            }
            QPushButton:hover { background: #252530; color: #FFF; }
        """)
        btn_close.clicked.connect(dialog.accept)
        
        btn_layout.addWidget(btn_open)
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)
        
        dialog.exec()


class GlobalSearchPanel(QWidget):
    """全局搜索面板 - 支持关键词+日期范围筛选"""
    
    def __init__(self, db: Database, parent=None):
        super().__init__(parent)
        self.db = db
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(20)
        
        # 标题区
        title = QLabel("全局搜索")
        title.setStyleSheet("font-size: 32px; font-weight: 800; color: #FFF; letter-spacing: 2px;")
        layout.addWidget(title)
        
        # 搜索输入
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("输入金额 (>100)、备注或日期...")
        self.search_input.setFixedHeight(50)
        self.search_input.setStyleSheet("""
            QLineEdit {
                background: #1C1C1E; border: 2px solid #333; border-radius: 12px;
                padding: 0 16px; color: #FFF; font-size: 16px;
            }
            QLineEdit:focus { border-color: #0A84FF; background: #222; }
        """)
        self.search_input.textChanged.connect(self._on_search)
        layout.addWidget(self.search_input)
        
        # 日期范围筛选栏
        date_style = """
            QDateEdit {
                background: #1C1C1E; color: #FFF; border: 1px solid #333;
                border-radius: 8px; padding: 8px 12px; font-size: 13px;
                min-width: 130px;
            }
            QDateEdit:focus { border-color: #FF9500; }
            QDateEdit::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: right center;
                width: 20px;
                border: none;
            }
            QDateEdit::down-arrow { image: none; }
        """
        
        date_row = QHBoxLayout()
        date_row.setSpacing(12)
        
        lbl_from = QLabel("从")
        lbl_from.setStyleSheet("color: #888; font-size: 13px; font-weight: 600;")
        date_row.addWidget(lbl_from)
        
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        self.date_from.setDisplayFormat("yyyy-MM-dd")
        self.date_from.setStyleSheet(date_style)
        self.date_from.dateChanged.connect(self._on_search)
        date_row.addWidget(self.date_from)
        
        lbl_to = QLabel("到")
        lbl_to.setStyleSheet("color: #888; font-size: 13px; font-weight: 600;")
        date_row.addWidget(lbl_to)
        
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setDisplayFormat("yyyy-MM-dd")
        self.date_to.setStyleSheet(date_style)
        self.date_to.dateChanged.connect(self._on_search)
        date_row.addWidget(self.date_to)
        
        # 日期筛选开关
        self.date_filter_btn = QPushButton("📅 日期筛选")
        self.date_filter_btn.setCheckable(True)
        self.date_filter_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.date_filter_btn.setStyleSheet("""
            QPushButton { 
                background: transparent; color: #888; border: 1px solid #333; 
                border-radius: 8px; padding: 8px 16px; font-weight: 600; font-size: 13px;
            }
            QPushButton:checked { background: #FF950020; color: #FF9500; border-color: #FF9500; }
            QPushButton:hover { border-color: #666; }
        """)
        self.date_filter_btn.clicked.connect(self._toggle_date_filter)
        date_row.addWidget(self.date_filter_btn)
        
        date_row.addStretch()
        layout.addLayout(date_row)
        
        # 初始隐藏日期选择器
        self.date_from.setVisible(False)
        self.date_to.setVisible(False)
        lbl_from.setVisible(False)
        lbl_to.setVisible(False)
        self._date_labels = (lbl_from, lbl_to)
        
        # 结果状态
        self.result_label = QLabel("输入关键词，搜索所有账本的交易记录。")
        self.result_label.setStyleSheet("color: #666; font-size: 14px; margin-left: 4px;")
        layout.addWidget(self.result_label)
        
        # 结果列表
        self.result_list = SmoothListWidget()
        self.result_list.setSpacing(6)
        self.result_list.setVerticalScrollMode(QListWidget.ScrollMode.ScrollPerPixel)
        layout.addWidget(self.result_list, 1)
    
    def _toggle_date_filter(self, checked: bool):
        self.date_from.setVisible(checked)
        self.date_to.setVisible(checked)
        self._date_labels[0].setVisible(checked)
        self._date_labels[1].setVisible(checked)
        self._on_search()
    
    def _on_search(self, *args):
        self.result_list.clear()
        keyword = self.search_input.text().strip()
        
        if not keyword and not self.date_filter_btn.isChecked():
            self.result_label.setText("等待输入...")
            return
        
        if keyword:
            results = self.db.search_transactions(keyword)
        else:
            # 无关键词但启用了日期筛选 → 获取全部记录
            self.db.cursor.execute("""
                SELECT t.id, t.amount, t.type, t.source, t.is_selected, t.created_at, m.name, m.icon
                FROM transactions t 
                JOIN modules m ON t.module_id = m.id
                ORDER BY t.created_at DESC
            """)
            results = self.db.cursor.fetchall()
        
        # 日期范围筛选
        if self.date_filter_btn.isChecked():
            date_from = self.date_from.date().toString("yyyy-MM-dd")
            date_to = self.date_to.date().toString("yyyy-MM-dd")
            filtered = []
            for trans in results:
                created_at = trans[5] if trans[5] else ""
                trans_date = created_at[:10]
                if date_from <= trans_date <= date_to:
                    filtered.append(trans)
            results = filtered
        
        self.result_label.setText(f"找到 {len(results)} 条匹配记录")
        
        for trans in results:
            trans_id, amount, trans_type, source, is_selected, created_at, mod_name, mod_icon = trans
            date_str = created_at[:10] if created_at else ""
            
            widget = TransactionItemWidget(
                trans_id, amount, trans_type, source,
                bool(is_selected), date_str, False, (mod_icon, mod_name)
            )
            item = QListWidgetItem()
            item.setSizeHint(QSize(0, 88))
            self.result_list.addItem(item)
            self.result_list.setItemWidget(item, widget)

    def refresh(self):
        keyword = self.search_input.text()
        if keyword.strip() or self.date_filter_btn.isChecked():
            self._on_search()


class MainWindow(QMainWindow):
    """
    World-Class Main Window
    Features: Frosted glass lock screen, premium sidebar layout, integrated search.
    """
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.current_module_name = ""
        self.current_module_id = None
        
        from .styles import THEME, MAIN_STYLESHEET
        self.theme = THEME
        
        self.setWindowTitle("Money Tracker Pro - Premium Edition")
        self.resize(1080, 720)
        self.setMinimumSize(900, 600)
        self.setStyleSheet(MAIN_STYLESHEET)
        
        # Load logic checks
        from common.security import SecurityManager
        self.security = SecurityManager()
        if self.security.has_password():
            self._show_lock_screen()
        else:
            self._init_main_ui()

    def _show_lock_screen(self):
        """World-class frosted glass lock screen."""
        self.lock_widget = QWidget()
        self.setCentralWidget(self.lock_widget)
        
        # Base background
        self.lock_widget.setStyleSheet(f"background-color: {self.theme['bg_void']};")
        main_layout = QVBoxLayout(self.lock_widget)
        main_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Glass card
        self.lock_card = QFrame()
        self.lock_card.setFixedSize(400, 480)
        self.lock_card.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #1A1A24, stop:1 #111118);
                border: 1px solid {self.theme['border_bright']};
                border-radius: 28px;
            }}
        """)
        apply_shadow(self.lock_card, blur=60, y=10, alpha=150)
        
        card_layout = QVBoxLayout(self.lock_card)
        card_layout.setContentsMargins(40, 50, 40, 50)
        card_layout.setSpacing(24)
        
        # App logo / icon
        logo = QLabel("🔒")
        logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo.setStyleSheet(f"""
            QLabel {{
                font-size: 56px;
                background: transparent;
                border: none;
            }}
        """)
        card_layout.addWidget(logo)
        
        card_layout.addSpacing(10)
        
        title = QLabel("欢迎回来")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet(f"font-size: 26px; font-weight: 800; color: {self.theme['text_p']}; background: transparent; border: none;")
        card_layout.addWidget(title)
        
        sub = QLabel("请输入密码解锁您的财务数据")
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sub.setStyleSheet(f"font-size: 13px; color: {self.theme['text_t']}; background: transparent; border: none;")
        card_layout.addWidget(sub)
        
        card_layout.addStretch()
        
        self.pwd_input = QLineEdit()
        self.pwd_input.setPlaceholderText("请输入密码")
        self.pwd_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.pwd_input.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.pwd_input.setStyleSheet(f"""
            QLineEdit {{
                background: {self.theme['bg_input']};
                border: 1px solid {self.theme['border']};
                border-radius: 14px;
                padding: 16px;
                font-size: 18px;
                letter-spacing: 4px;
                color: {self.theme['text_p']};
            }}
            QLineEdit:focus {{
                border: 1.5px solid {self.theme['accent']}88;
                background: {self.theme['bg_layer2']};
            }}
        """)
        self.pwd_input.returnPressed.connect(self._verify_password)
        card_layout.addWidget(self.pwd_input)
        
        unlock_btn = QPushButton("解 锁")
        unlock_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        unlock_btn.setFixedHeight(54)
        unlock_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {self.theme['accent']}, stop:1 {self.theme['accent_deep']});
                color: #000;
                border: none;
                border-radius: 14px;
                font-size: 16px;
                font-weight: 800;
                letter-spacing: 4px;
            }}
            QPushButton:hover {{ opacity: 0.9; }}
        """)
        unlock_btn.clicked.connect(self._verify_password)
        card_layout.addWidget(unlock_btn)
        
        main_layout.addWidget(self.lock_card)
        self.pwd_input.setFocus()
        
    def _verify_password(self):
        pwd = self.pwd_input.text()
        if self.security.verify_password(pwd):
            from .effects import fade_in
            QTimer.singleShot(50, self._perform_unlock) # Use QTimer to detach from signal handler
        else:
            self.pwd_input.clear()
            from .effects import shake_widget
            shake_widget(self.lock_card)
            self.pwd_input.setPlaceholderText("密码错误，请重试")
            self.pwd_input.setStyleSheet(f"""
                QLineEdit {{
                    background: rgba(255,69,58,0.08);
                    border: 1px solid #FF453A88;
                    border-radius: 14px;
                    padding: 16px;
                    font-size: 18px;
                    color: {self.theme['text_p']};
                }}
            """)
            
    def _perform_unlock(self):
        """延迟执行解锁界面切换，防止事件循环冲突导致的闪退"""
        if hasattr(self, 'lock_widget'):
            self.lock_widget.hide()
        self._init_main_ui()
        from .effects import fade_in
        fade_in(self, duration=400)

    # --- 主界面 ---
    def _init_main_ui(self):
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        main_layout = QHBoxLayout(self.central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # ── Sidebar ───────────────────────────────────────────────
        left_panel = QFrame()
        left_panel.setFixedWidth(260)
        left_panel.setStyleSheet(f"""
            QFrame {{
                background-color: {self.theme['bg_deep']};
                border-right: 1px solid {self.theme['border']};
            }}
        """)
        
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(0)
        self._setup_sidebar(left_layout)
        main_layout.addWidget(left_panel)
        
        # ── Content Stack ─────────────────────────────────────────
        self.content_stack = QStackedWidget()
        self.content_stack.setStyleSheet(f"background-color: {self.theme['bg_void']};")
        
        self.module_panel = ModuleContentPanel(self.db)
        self.search_panel = GlobalSearchPanel(self.db)
        self.summary_panel = SummaryPanel()
        
        self.content_stack.addWidget(self.module_panel)
        self.content_stack.addWidget(self.search_panel)
        self.content_stack.addWidget(self.summary_panel)
        
        main_layout.addWidget(self.content_stack, 1)
        
        # Wire real-time analytics refresh: when any transaction changes, refresh sidebar & summary
        self.module_panel.data_changed.connect(self._on_data_changed)
        
        # Load data
        self._load_modules()
        self._update_shortcuts()

        
    def _setup_sidebar(self, layout):
        # Premium Brand Logo Section
        logo_area = QWidget()
        logo_area.setFixedHeight(120)
        logo_layout = QHBoxLayout(logo_area)
        logo_layout.setContentsMargins(28, 40, 24, 20)
        logo_layout.setSpacing(12)
        
        logo = QLabel("∞")
        logo.setStyleSheet(f"""
            font-size: 38px; font-weight: 900; 
            color: {self.theme['accent']};
            background: transparent; border: none;
        """)
        
        brand_vbox = QVBoxLayout()
        brand_vbox.setSpacing(0)
        brand_vbox.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        
        brand_main = QLabel("智能财务")
        brand_main.setStyleSheet(f"font-size: 20px; font-weight: 900; color: {self.theme['text_p']}; letter-spacing: 2px;")
        brand_sub = QLabel("记账管家专业版")
        brand_sub.setStyleSheet(f"font-size: 10px; font-weight: 700; color: {self.theme['accent']}; letter-spacing: 2px;")
        
        brand_vbox.addWidget(brand_main)
        brand_vbox.addWidget(brand_sub)
        
        logo_layout.addWidget(logo)
        logo_layout.addLayout(brand_vbox)
        logo_layout.addStretch()
        layout.addWidget(logo_area)
        
        # Header for lists
        nav_header = QHBoxLayout()
        nav_header.setContentsMargins(28, 10, 20, 10)
        lbl_ledgers = QLabel("账本列表")
        lbl_ledgers.setStyleSheet("font-size: 11px; font-weight: 700; color: #8A8A9E; letter-spacing: 1px;")
        nav_header.addWidget(lbl_ledgers)
        nav_header.addStretch()
        layout.addLayout(nav_header)

        # ── Sort selector ─────────────────────────────────────────
        from PyQt6.QtWidgets import QComboBox
        sort_row = QHBoxLayout()
        sort_row.setContentsMargins(16, 2, 16, 4)

        sort_lbl = QLabel("排序：")
        sort_lbl.setStyleSheet(f"font-size:12px; color:{self.theme['text_t']}; background:transparent;")
        sort_row.addWidget(sort_lbl)

        self.sort_combo = QComboBox()
        self.sort_combo.addItems(["余额从大到小", "余额从小到大", "创建时间", "最近动态"])
        self.sort_combo.setStyleSheet(f"""
            QComboBox {{
                background: {self.theme['bg_layer2']};
                color: {self.theme['text_s']};
                border: 1px solid {self.theme['border']};
                border-radius: 8px;
                padding: 3px 10px;
                font-size: 12px;
            }}
            QComboBox::drop-down {{ border: none; width: 18px; }}
            QComboBox QAbstractItemView {{
                background: {self.theme['bg_layer2']};
                color: {self.theme['text_p']};
                border: 1px solid {self.theme['border_bright']};
                selection-background-color: {self.theme['accent']}33;
            }}
        """)
        _SORT_KEYS = ["balance_desc", "balance_asc", "created_desc", "updated_desc"]
        def _on_sort_changed(idx):
            prev_mid = None
            cur = self.module_list.currentItem()
            if cur:
                prev_mid = cur.data(Qt.ItemDataRole.UserRole)[0]
            self._load_modules(_SORT_KEYS[idx])
            # Restore previous selection if still present
            if prev_mid:
                for i in range(self.module_list.count()):
                    if self.module_list.item(i).data(Qt.ItemDataRole.UserRole)[0] == prev_mid:
                        self.module_list.setCurrentRow(i)
                        break
        self.sort_combo.currentIndexChanged.connect(_on_sort_changed)
        sort_row.addWidget(self.sort_combo, 1)
        layout.addLayout(sort_row)
        
        # List of ledgers
        list_container = QWidget()
        lc_layout = QVBoxLayout(list_container)
        lc_layout.setContentsMargins(14, 0, 14, 0)
        
        self.module_list = SmoothListWidget()
        self.module_list.setCursor(Qt.CursorShape.PointingHandCursor)
        self.module_list.currentRowChanged.connect(self._on_module_changed)
        self.module_list.setStyleSheet(f"""
            QListWidget {{ background: transparent; border: none; }}
            QListWidget::item {{
                height: 44px;
                border-radius: 12px;
                padding-left: 14px;
                margin-bottom: 4px;
                color: {self.theme['text_s']};
                font-size: 14px;
                font-weight: 600;
            }}
            QListWidget::item:hover {{
                background: rgba(255,255,255,0.06);
                color: {self.theme['text_p']};
            }}
            QListWidget::item:selected {{
                background: rgba(255,184,0,0.15);
                color: {self.theme['accent']};
            }}
        """)
        self.module_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.module_list.customContextMenuRequested.connect(self._show_module_context_menu)
        lc_layout.addWidget(self.module_list)
        layout.addWidget(list_container, 1)
        
        # Add Ledger button
        add_btn_layout = QHBoxLayout()
        add_btn_layout.setContentsMargins(28, 10, 28, 20)
        
        from .effects import RippleButton
        add_btn = RippleButton("＋ 新建账本")
        add_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        add_btn.setFixedHeight(44)
        add_btn.clicked.connect(self._add_module)
        add_btn.setStyleSheet(f"""
            QPushButton {{
                background: transparent; color: {self.theme['text_s']};
                border: 1px dashed {self.theme['border_bright']};
                border-radius: 12px; font-weight: 700; font-size: 13px;
            }}
            QPushButton:hover {{ background: rgba(255,255,255,0.05); color: {self.theme['text_p']}; border-color: rgba(255,255,255,0.3); }}
        """)
        add_btn_layout.addWidget(add_btn, 1)
        layout.addLayout(add_btn_layout)
        
        # Bottom tools (Search, Stats, Settings)
        bottom_area = QWidget()
        bottom_layout = QVBoxLayout(bottom_area)
        bottom_layout.setContentsMargins(14, 20, 14, 24)
        bottom_layout.setSpacing(6)
        
        def _nav_btn(text, icon, callback):
            btn = RippleButton(f"{icon}  {text}")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setFixedHeight(46)
            btn.setStyleSheet(f"""
                QPushButton {{
                    text-align: left; padding-left: 16px;
                    border: none; border-radius: 12px;
                    background: transparent; color: {self.theme['text_s']};
                    font-size: 14px; font-weight: 600;
                }}
                QPushButton:hover {{
                    background: rgba(255,255,255,0.06); color: {self.theme['text_p']};
                }}
            """)
            btn.clicked.connect(callback)
            return btn
            
        btn_search = _nav_btn("全局搜索", "🔍", lambda: self._show_panel("search"))
        btn_stats = _nav_btn("数据分析", "📊", lambda: self._show_panel("summary"))
        btn_settings = _nav_btn("系统设置", "⚙️", self._open_settings)
        
        bottom_layout.addWidget(btn_search)
        bottom_layout.addWidget(btn_stats)
        bottom_layout.addWidget(btn_settings)
        
        # Divider
        div = QFrame()
        div.setFixedHeight(1)
        div.setStyleSheet(f"background: {self.theme['border']}; border: none; margin: 0 14px;")
        layout.addWidget(div)
        layout.addWidget(bottom_area)

    def _load_modules(self, sort_by: str = None):
        # Persist sort choice across reloads
        if sort_by is not None:
            self._current_sort = sort_by
        sort = getattr(self, '_current_sort', 'balance_desc')

        if sort == 'balance_asc':
            module_list = self.db.get_modules_sorted_by_balance_asc()
        elif sort == 'created_desc':
            module_list = self.db.get_modules_sorted_by_created()
        elif sort == 'updated_desc':
            module_list = self.db.get_modules_sorted_by_updated()
        else:  # default: balance_desc
            module_list = self.db.get_modules_sorted_by_balance()

        self.module_list.clear()
        self.module_data_full = module_list
        detailed_modules = []
        for mid, name, icon, color, balance in module_list:
            inc, exp = self.db.get_module_stats(mid)
            detailed_modules.append((mid, name, icon, color, inc, exp))
            item = QListWidgetItem(f"{icon}  {name}")
            item.setData(Qt.ItemDataRole.UserRole, (mid, name, icon, color))
            item.setToolTip(f"精确余额: ¥{balance:,.2f}")
            self.module_list.addItem(item)

        self.module_data = detailed_modules

        # Always keep the analytics (summary) panel current
        self._refresh_summary()

        if self.module_list.count() > 0:
            self.module_list.setCurrentRow(0)

    def _refresh_summary(self):
        """Push the latest stats to the analytics panel (real-time update)."""
        if not hasattr(self, 'module_data') or not hasattr(self, 'summary_panel'):
            return
        total_income  = sum(m[4] for m in self.module_data)
        total_expense = sum(m[5] for m in self.module_data)
        total_balance = total_income - total_expense
        self.summary_panel.update_data(total_balance, total_income, total_expense, self.module_data)

    def _on_data_changed(self):
        """Called when transactions are added/deleted — refreshes sidebar balances + analytics.
        Preserves the current module selection so the ledger view stays in place."""
        # Remember which row was selected
        current_row = self.module_list.currentRow()

        sort = getattr(self, '_current_sort', 'balance_desc')
        if sort == 'balance_asc':
            module_list = self.db.get_modules_sorted_by_balance_asc()
        elif sort == 'created_desc':
            module_list = self.db.get_modules_sorted_by_created()
        elif sort == 'updated_desc':
            module_list = self.db.get_modules_sorted_by_updated()
        else:
            module_list = self.db.get_modules_sorted_by_balance()

        self.module_list.blockSignals(True)
        self.module_list.clear()
        detailed_modules = []
        for mid, name, icon, color, balance in module_list:
            inc, exp = self.db.get_module_stats(mid)
            detailed_modules.append((mid, name, icon, color, inc, exp))
            item = QListWidgetItem(f"{icon}  {name}")
            item.setData(Qt.ItemDataRole.UserRole, (mid, name, icon, color))
            item.setToolTip(f"精确余额: ¥{balance:,.2f}")
            self.module_list.addItem(item)

        self.module_data = detailed_modules

        # Restore selection without triggering full reload
        if 0 <= current_row < self.module_list.count():
            self.module_list.setCurrentRow(current_row)
        self.module_list.blockSignals(False)

        # Push fresh data to analytics
        self._refresh_summary()

            
    def _on_module_changed(self, row: int):
        if row < 0: return
        item = self.module_list.item(row)
        if item:
            mid, name, icon, color = item.data(Qt.ItemDataRole.UserRole)
            self.current_module_id = mid
            self.current_module_name = name
            # Navigate to module panel
            self.module_panel.load_module(mid, name, icon)
            self.content_stack.setCurrentWidget(self.module_panel)

    def _show_panel(self, panel: str):
        """Switch the content area to the named panel."""
        if panel == "search":
            self.content_stack.setCurrentWidget(self.search_panel)
            self.search_panel.refresh()
        elif panel == "summary":
            self.content_stack.setCurrentWidget(self.summary_panel)
            # Compute needed values for summary_panel.update_data(total_balance, total_income, total_expense, modules_data)
            if hasattr(self, 'module_data'):
                total_income  = sum(m[4] for m in self.module_data)
                total_expense = sum(m[5] for m in self.module_data)
                total_balance = total_income - total_expense
                self.summary_panel.update_data(total_balance, total_income, total_expense, self.module_data)

    def _add_module(self):
        """Open the dialog to create a new ledger."""
        dialog = ModuleDialog(self, db=self.db)
        if dialog.exec() == 1:
            name, icon, color = dialog.get_result()
            self.db.add_module(name, icon, color)
            self._load_modules()
            self.module_list.setCurrentRow(self.module_list.count() - 1)

    def _edit_current_module(self):
        item = self.module_list.currentItem()
        if not item: return
        
        mid, name, icon, color = item.data(Qt.ItemDataRole.UserRole)
        dialog = ModuleDialog(self, (mid, name, icon, color), db=self.db)
        result = dialog.exec()
        
        if result == 1:
            new_name, new_icon, new_color = dialog.get_result()
            self.db.update_module(mid, new_name, new_icon, new_color)
            self._load_modules()
            # Restore selection
            for i in range(self.module_list.count()):
                if self.module_list.item(i).data(Qt.ItemDataRole.UserRole)[0] == mid:
                    self.module_list.setCurrentRow(i)
                    break
        elif result == 2:
            self.db.delete_module(mid)
            self._load_modules()
    def _show_module_context_menu(self, pos):
        item = self.module_list.itemAt(pos)
        if not item: return
        # Select the right-clicked item
        self.module_list.setCurrentItem(item)
        
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background-color: #2C2C2E; border: 1px solid #48484A; border-radius: 8px; padding: 4px; }
            QMenu::item { background-color: transparent; color: #FFFFFF; padding: 8px 16px; border-radius: 4px; }
            QMenu::item:selected { background-color: #0A84FF; }
        """)
        
        edit_action = menu.addAction("✏️ 编辑")
        delete_action = menu.addAction("🗑️ 删除")
        
        action = menu.exec(self.module_list.mapToGlobal(pos))
        
        if action == edit_action:
            self._edit_current_module()
        elif action == delete_action:
            self._delete_current_module()

    def _delete_current_module(self):
        """Delete the currently selected ledger after confirmation."""
        item = self.module_list.currentItem()
        if not item: return
        
        mid, name, icon, color = item.data(Qt.ItemDataRole.UserRole)
        from PyQt6.QtWidgets import QMessageBox
        reply = QMessageBox.question(
            self, "确认删除",
            f"确定要删除账本「{name}」吗？\n所有交易记录将永久删除。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete_module(mid)
            self._load_modules()

    def _open_settings(self):
        from .settings_dialog import SettingsDialog
        dialog = SettingsDialog(self.security, self)
        if dialog.exec():
            # Reload shortcuts in case the user changed key bindings
            self._update_shortcuts()
            
    def _update_shortcuts(self):
        from PyQt6.QtCore import QSettings
        settings = QSettings("MoneyTracker", "Shortcuts")
        key_income  = settings.value("key_income",  "Z")
        key_expense = settings.value("key_expense", "X")
        
        # Clear previous shortcuts if they exist
        if hasattr(self, '_shortcut_income'):
            self._shortcut_income.setEnabled(False)
            self._shortcut_expense.setEnabled(False)
        
        self._shortcut_income = QShortcut(QKeySequence(key_income), self)
        self._shortcut_income.activated.connect(self._trigger_income)
        
        self._shortcut_expense = QShortcut(QKeySequence(key_expense), self)
        self._shortcut_expense.activated.connect(self._trigger_expense)
        
    def closeEvent(self, event):
        # Allow animations to finish before closing
        event.accept()

    def _trigger_income(self):
        if hasattr(self, 'module_panel') and self.content_stack.currentWidget() == self.module_panel:
            self.module_panel._add_transaction("income")
        else:
            if not self.module_list.count(): return
            item = self.module_list.item(0)
            mid, name, icon, color = item.data(Qt.ItemDataRole.UserRole)
            from .transaction_dialog import TransactionDialog
            dialog = TransactionDialog("income", name, self)
            if dialog.exec():
                amt, src = dialog.get_result()
                self.db.add_transaction(amt, "income", src, mid)
                self._load_modules()

    def _trigger_expense(self):
        if hasattr(self, 'module_panel') and self.content_stack.currentWidget() == self.module_panel:
            self.module_panel._add_transaction("expense")
        else:
            if not self.module_list.count(): return
            item = self.module_list.item(0)
            mid, name, icon, color = item.data(Qt.ItemDataRole.UserRole)
            from .transaction_dialog import TransactionDialog
            dialog = TransactionDialog("expense", name, self)
            if dialog.exec():
                amt, src = dialog.get_result()
                self.db.add_transaction(amt, "expense", src, mid)
                self._load_modules()